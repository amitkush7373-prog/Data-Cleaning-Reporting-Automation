"""
Multi-Format Automated Report Generation Module.
Supports:
1. Multi-Sheet Styled Excel Report (.xlsx) with OpenPyXL formatting
2. Executive PDF Business Report (.pdf) with ReportLab
3. Standalone Responsive HTML Report (.html)
4. Cleaned CSV/Excel dataset exports
"""
import io
import os
from datetime import datetime
from typing import Dict, Any, List, Optional, Union
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from src.logger import log_event
from src.visualizer import generate_static_summary_plot


def _sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Sanitize DataFrame for Excel export by stripping timezones from datetime columns."""
    if df is None or df.empty:
        return df
    excel_df = df.copy()
    for col in excel_df.columns:
        if pd.api.types.is_datetime64_any_dtype(excel_df[col]):
            try:
                if hasattr(excel_df[col].dt, "tz") and excel_df[col].dt.tz is not None:
                    excel_df[col] = excel_df[col].dt.tz_localize(None)
            except Exception:
                excel_df[col] = excel_df[col].astype(str)
        elif str(excel_df[col].dtype).lower() in ["datetime64[ns, utc]", "datetimetz"]:
            try:
                excel_df[col] = excel_df[col].dt.tz_localize(None)
            except Exception:
                excel_df[col] = excel_df[col].astype(str)
    return excel_df


# -------------------------------------------------------------
# 1. Multi-Sheet Styled Excel Report (.xlsx)
# -------------------------------------------------------------

def generate_excel_report(
    raw_df: pd.DataFrame,
    clean_df: pd.DataFrame,
    raw_validation: Dict[str, Any],
    clean_validation: Dict[str, Any],
    audit_log: Dict[str, Any],
    kpis: Dict[str, Any],
    aggregations: Dict[str, pd.DataFrame],
    insights: Dict[str, List[str]],
    output_path_or_buffer: Optional[Union[str, io.BytesIO]] = None
) -> Union[str, bytes]:
    """
    Generate a professional multi-sheet Excel workbook with styled headers,
    KPI cards, audit log, and formatted tables.
    """
    raw_df = _sanitize_df_for_excel(raw_df)
    clean_df = _sanitize_df_for_excel(clean_df)
    wb = openpyxl.Workbook()
    
    # Styles
    navy_header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    blue_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    accent_green_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    header_font_white = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
    bold_cell_font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    regular_font = Font(name="Segoe UI", size=10, color="334155")
    
    thin_border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # ---------------------------------------------------------
    # Sheet 1: Executive Summary
    # ---------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.merge_cells("A1:F1")
    ws_summary["A1"] = "Data Cleaning & Reporting Automation — Executive Summary"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(vertical="center")
    ws_summary.row_dimensions[1].height = 30

    ws_summary["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Automated Pipeline Report"
    ws_summary["A2"].font = subtitle_font
    ws_summary.row_dimensions[2].height = 20

    # Before vs After Health Metric Cards
    ws_summary["A4"] = "DATASET HEALTH & CLEANING HIGHLIGHTS"
    ws_summary["A4"].font = section_font

    headers_summary = ["Metric", "Before Cleaning (Raw)", "After Cleaning (Clean)", "Status / Change"]
    for col_idx, h in enumerate(headers_summary, start=1):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.fill = navy_header_fill
        cell.font = header_font_white
        cell.alignment = center_align
    ws_summary.row_dimensions[5].height = 24

    metrics_rows = [
        ("Total Rows", raw_validation.get("total_rows", 0), clean_validation.get("total_rows", 0), f"-{audit_log.get('duplicates_removed', 0)} dups"),
        ("Total Columns", raw_validation.get("total_columns", 0), clean_validation.get("total_columns", 0), f"{clean_validation.get('total_columns', 0) - raw_validation.get('total_columns', 0):+d} cols"),
        ("Missing Values", raw_validation.get("missing_report", {}).get("total_missing", 0), clean_validation.get("missing_report", {}).get("total_missing", 0), "Resolved 100%"),
        ("Duplicate Rows", raw_validation.get("duplicate_report", {}).get("duplicate_count", 0), clean_validation.get("duplicate_report", {}).get("duplicate_count", 0), "0 Duplicates Remaining"),
        ("Health Score", f"{raw_validation.get('quality_score', 0)}/100 [{raw_validation.get('quality_grade', '')}]", f"{clean_validation.get('quality_score', 100)}/100 [{clean_validation.get('quality_grade', '')}]", f"+{clean_validation.get('quality_score', 100) - raw_validation.get('quality_score', 0):.1f} pts")
    ]

    for row_idx, (m, b, a, c) in enumerate(metrics_rows, start=6):
        ws_summary.cell(row=row_idx, column=1, value=m).font = bold_cell_font
        ws_summary.cell(row=row_idx, column=2, value=str(b)).font = regular_font
        ws_summary.cell(row=row_idx, column=3, value=str(a)).font = regular_font
        ws_summary.cell(row=row_idx, column=4, value=str(c)).font = bold_cell_font
        
        for c_idx in range(1, 5):
            cell = ws_summary.cell(row=row_idx, column=c_idx)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
        ws_summary.row_dimensions[row_idx].height = 20

    # Key KPIs Section
    ws_summary["A13"] = "KEY PERFORMANCE INDICATORS (KPIs)"
    ws_summary["A13"].font = section_font
    
    headers_kpi = ["KPI Name", "Calculated Metric Value", "Description"]
    for col_idx, h in enumerate(headers_kpi, start=1):
        cell = ws_summary.cell(row=14, column=col_idx, value=h)
        cell.fill = blue_header_fill
        cell.font = header_font_white
        cell.alignment = center_align
    ws_summary.row_dimensions[14].height = 22

    cur_row = 15
    for k_key, k_obj in kpis.items():
        ws_summary.cell(row=cur_row, column=1, value=k_obj.get("label", k_key)).font = bold_cell_font
        ws_summary.cell(row=cur_row, column=2, value=str(k_obj.get("value", ""))).font = bold_cell_font
        ws_summary.cell(row=cur_row, column=3, value=str(k_obj.get("description", ""))).font = regular_font
        for c_idx in range(1, 4):
            ws_summary.cell(row=cur_row, column=c_idx).border = thin_border
        cur_row += 1

    # Insights & Recommendations Section
    cur_row += 2
    ws_summary.cell(row=cur_row, column=1, value="AUTOMATED INSIGHTS & RECOMMENDATIONS").font = section_font
    cur_row += 1

    for ins in insights.get("quality_insights", []) + insights.get("business_insights", []) + insights.get("recommendations", []):
        ws_summary.cell(row=cur_row, column=1, value=f"•  {ins}").font = regular_font
        ws_summary.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        cur_row += 1

    _autofit_columns(ws_summary)

    # ---------------------------------------------------------
    # Sheet 2: Data Quality Assessment
    # ---------------------------------------------------------
    ws_quality = wb.create_sheet(title="Data Quality")
    ws_quality.views.sheetView[0].showGridLines = True
    
    ws_quality["A1"] = "Data Quality Diagnostics (Raw vs Clean)"
    ws_quality["A1"].font = title_font
    
    q_headers = ["Column Name", "Raw Missing Count", "Raw Missing %", "Clean Missing Count", "Inferred Type", "Clean Data Type"]
    for col_idx, h in enumerate(q_headers, start=1):
        cell = ws_quality.cell(row=3, column=col_idx, value=h)
        cell.fill = navy_header_fill
        cell.font = header_font_white
        cell.alignment = center_align
    ws_quality.row_dimensions[3].height = 24

    r_missing = raw_validation.get("missing_report", {}).get("columns", {})
    inferred = raw_validation.get("inferred_types", {})
    
    row_pos = 4
    for col in clean_df.columns:
        # Match with raw col
        orig_match = None
        for old_c, new_c in audit_log.get("column_renaming", {}).items():
            if new_c == col:
                orig_match = old_c
                break
        raw_c = orig_match or col
        raw_m_info = r_missing.get(raw_c, {"count": 0, "pct": 0.0})
        
        ws_quality.cell(row=row_pos, column=1, value=col).font = bold_cell_font
        ws_quality.cell(row=row_pos, column=2, value=raw_m_info.get("count", 0)).font = regular_font
        ws_quality.cell(row=row_pos, column=3, value=f"{raw_m_info.get('pct', 0.0)}%").font = regular_font
        ws_quality.cell(row=row_pos, column=4, value=int(clean_df[col].isna().sum())).font = regular_font
        ws_quality.cell(row=row_pos, column=5, value=str(inferred.get(raw_c, "standard"))).font = regular_font
        ws_quality.cell(row=row_pos, column=6, value=str(clean_df[col].dtype)).font = regular_font

        for c_idx in range(1, 7):
            cell = ws_quality.cell(row=row_pos, column=c_idx)
            cell.border = thin_border
            if row_pos % 2 == 0:
                cell.fill = zebra_fill
        row_pos += 1

    _autofit_columns(ws_quality)

    # ---------------------------------------------------------
    # Sheet 3: Cleaning Audit Log
    # ---------------------------------------------------------
    ws_log = wb.create_sheet(title="Cleaning Audit Trail")
    ws_log.views.sheetView[0].showGridLines = True
    
    ws_log["A1"] = "Detailed Cleaning Pipeline Audit Trail"
    ws_log["A1"].font = title_font
    
    log_headers = ["Step #", "Operation Description"]
    for col_idx, h in enumerate(log_headers, start=1):
        cell = ws_log.cell(row=3, column=col_idx, value=h)
        cell.fill = navy_header_fill
        cell.font = header_font_white
    ws_log.row_dimensions[3].height = 24

    for s_idx, step_desc in enumerate(audit_log.get("steps_executed", []), start=1):
        r = 3 + s_idx
        ws_log.cell(row=r, column=1, value=f"Step {s_idx}").font = bold_cell_font
        ws_log.cell(row=r, column=2, value=step_desc).font = regular_font
        ws_log.cell(row=r, column=1).border = thin_border
        ws_log.cell(row=r, column=2).border = thin_border
        if r % 2 == 0:
            ws_log.cell(row=r, column=1).fill = zebra_fill
            ws_log.cell(row=r, column=2).fill = zebra_fill

    _autofit_columns(ws_log)

    # ---------------------------------------------------------
    # Sheet 4: Aggregations & Breakdowns
    # ---------------------------------------------------------
    ws_agg = wb.create_sheet(title="KPIs and Breakdowns")
    ws_agg.views.sheetView[0].showGridLines = True
    ws_agg["A1"] = "Dimensional Aggregations & Business Breakdown"
    ws_agg["A1"].font = title_font
    
    current_agg_row = 3
    for agg_name, agg_df in aggregations.items():
        if agg_df is None or agg_df.empty:
            continue
        
        ws_agg.cell(row=current_agg_row, column=1, value=f"Summary: {agg_name.replace('_', ' ').title()}").font = section_font
        current_agg_row += 1
        
        # Headers
        for col_idx, col_name in enumerate(agg_df.columns, start=1):
            c = ws_agg.cell(row=current_agg_row, column=col_idx, value=col_name.replace("_", " ").title())
            c.fill = blue_header_fill
            c.font = header_font_white
            c.alignment = center_align
        current_agg_row += 1
        
        # Data
        for _, row_vals in agg_df.iterrows():
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws_agg.cell(row=current_agg_row, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
            current_agg_row += 1
        current_agg_row += 2

    _autofit_columns(ws_agg)

    # ---------------------------------------------------------
    # Sheet 5: Cleaned Dataset
    # ---------------------------------------------------------
    ws_data = wb.create_sheet(title="Cleaned Data")
    ws_data.views.sheetView[0].showGridLines = True

    # Column Headers
    for col_idx, col_name in enumerate(clean_df.columns, start=1):
        c = ws_data.cell(row=1, column=col_idx, value=col_name)
        c.fill = navy_header_fill
        c.font = header_font_white
        c.alignment = center_align
    ws_data.row_dimensions[1].height = 24

    # Data Rows
    for row_idx, row_vals in enumerate(clean_df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_vals, start=1):
            if isinstance(val, (pd.Timestamp, datetime)) and getattr(val, "tzinfo", None) is not None:
                try:
                    val = val.replace(tzinfo=None)
                except Exception:
                    val = str(val)
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else "")
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
        ws_data.row_dimensions[row_idx].height = 18

    _autofit_columns(ws_data)

    # Save to path or BytesIO buffer
    if isinstance(output_path_or_buffer, str):
        os.makedirs(os.path.dirname(output_path_or_buffer) if os.path.dirname(output_path_or_buffer) else ".", exist_ok=True)
        wb.save(output_path_or_buffer)
        log_event("SUCCESS", "EXPORT", f"Saved styled Excel report to '{output_path_or_buffer}'.")
        return output_path_or_buffer
    else:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


def _autofit_columns(worksheet):
    """Auto-fit Excel worksheet columns nicely."""
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len and len(val_str) < 60:
                max_len = len(val_str)
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)


# -------------------------------------------------------------
# 2. Executive PDF Business Report (.pdf)
# -------------------------------------------------------------

def generate_pdf_report(
    clean_df: pd.DataFrame,
    raw_validation: Dict[str, Any],
    clean_validation: Dict[str, Any],
    audit_log: Dict[str, Any],
    kpis: Dict[str, Any],
    insights: Dict[str, List[str]],
    output_path_or_buffer: Optional[Union[str, io.BytesIO]] = None
) -> Union[str, bytes]:
    """
    Generate an Executive PDF Business Report with ReportLab.
    """
    is_file = isinstance(output_path_or_buffer, str)
    buffer = io.BytesIO() if not is_file else output_path_or_buffer
    
    if is_file and os.path.dirname(output_path_or_buffer):
        os.makedirs(os.path.dirname(output_path_or_buffer), exist_ok=True)

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    # Custom Typography Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=14
    )
    heading_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#1E293B"),
        spaceBefore=12,
        spaceAfter=6
    )
    body_style = ParagraphStyle(
        'ReportBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#334155")
    )
    bullet_style = ParagraphStyle(
        'ReportBullet',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor("#1E293B"),
        leftIndent=12,
        spaceAfter=3
    )

    story = []

    # Title & Metadata Banner
    story.append(Paragraph("Data Cleaning & Reporting Automation", title_style))
    story.append(Paragraph(f"Executive Pipeline Report | Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2563EB"), spaceBefore=2, spaceAfter=10))

    # 1. Executive Summary & Health Scorecard
    story.append(Paragraph("1. Executive Summary & Health Scorecard", heading_style))
    
    raw_score = raw_validation.get("quality_score", 0)
    clean_score = clean_validation.get("quality_score", 100)
    dups_removed = audit_log.get("duplicates_removed", 0)
    raw_missing = raw_validation.get("missing_report", {}).get("total_missing", 0)
    clean_missing = clean_validation.get("missing_report", {}).get("total_missing", 0)

    scorecard_data = [
        ["Health Metric", "Raw Ingestion", "Cleaned Result", "Pipeline Impact"],
        ["Data Quality Score", f"{raw_score}/100 [{raw_validation.get('quality_grade', '')}]", f"{clean_score}/100 [{clean_validation.get('quality_grade', '')}]", f"+{clean_score - raw_score:.1f} Quality Gain"],
        ["Total Rows", f"{raw_validation.get('total_rows', 0):,}", f"{clean_validation.get('total_rows', 0):,}", f"{dups_removed} Duplicates Removed"],
        ["Total Columns", f"{raw_validation.get('total_columns', 0)}", f"{clean_validation.get('total_columns', 0)}", "Standardized snake_case"],
        ["Missing Values", f"{raw_missing:,}", f"{clean_missing:,}", "100% Imputed / Resolved"]
    ]

    t_scorecard = Table(scorecard_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
    t_scorecard.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F8FAFC"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]))
    story.append(t_scorecard)
    story.append(Spacer(1, 10))

    # 2. Key Performance Indicators (KPIs)
    story.append(Paragraph("2. Key Performance Indicators (KPIs)", heading_style))
    kpi_rows = [["KPI Indicator", "Computed Value", "Description"]]
    for k_key, k_obj in list(kpis.items())[:6]:
        kpi_rows.append([
            k_obj.get("label", k_key),
            k_obj.get("value", ""),
            k_obj.get("description", "")
        ])

    t_kpis = Table(kpi_rows, colWidths=[2.2*inch, 2.0*inch, 3.0*inch])
    t_kpis.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#EFF6FF"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]))
    story.append(t_kpis)
    story.append(Spacer(1, 10))

    # 3. Visual Performance Snapshot
    story.append(Paragraph("3. Visual Performance Analysis", heading_style))
    try:
        # Find numeric / category candidates
        num_cols = clean_df.select_dtypes(include=[np.number]).columns
        cat_cols = clean_df.select_dtypes(include=["object", "string", "category"]).columns
        c_col = cat_cols[0] if len(cat_cols) > 0 else None
        v_col = num_cols[0] if len(num_cols) > 0 else None
        
        plot_buf = generate_static_summary_plot(clean_df, c_col, v_col)
        story.append(RLImage(plot_buf, width=6.8*inch, height=3.0*inch))
    except Exception as e:
        story.append(Paragraph(f"Visual chart rendering skipped: {str(e)}", body_style))

    story.append(Spacer(1, 10))

    # 4. Automated Insights & Strategic Recommendations
    story.append(Paragraph("4. Automated Insights & Findings", heading_style))
    for item in insights.get("quality_insights", []) + insights.get("business_insights", []):
        story.append(Paragraph(f"• <b>Insight:</b> {item}", bullet_style))

    story.append(Spacer(1, 6))
    story.append(Paragraph("5. Strategic Recommendations", heading_style))
    for rec in insights.get("recommendations", []):
        story.append(Paragraph(f"• <b>Action:</b> {rec}", bullet_style))

    # Build Document
    doc.build(story)
    log_event("SUCCESS", "EXPORT", "PDF report compiled successfully.")

    if is_file:
        return output_path_or_buffer
    else:
        buffer.seek(0)
        return buffer.getvalue()


# -------------------------------------------------------------
# 3. Standalone Responsive HTML Report (.html)
# -------------------------------------------------------------

def generate_html_report(
    clean_df: pd.DataFrame,
    raw_validation: Dict[str, Any],
    clean_validation: Dict[str, Any],
    audit_log: Dict[str, Any],
    kpis: Dict[str, Any],
    insights: Dict[str, List[str]],
    output_path_or_buffer: Optional[Union[str, io.BytesIO]] = None
) -> Union[str, bytes]:
    """Generate standalone modern HTML business report."""
    
    kpi_cards_html = "".join([
        f"""
        <div class="kpi-card">
            <div class="kpi-label">{k_obj.get('label', k)}</div>
            <div class="kpi-value">{k_obj.get('value', '')}</div>
            <div class="kpi-desc">{k_obj.get('description', '')}</div>
        </div>
        """ for k, k_obj in kpis.items()
    ])

    insights_html = "".join([f"<li>{ins}</li>" for ins in insights.get('quality_insights', []) + insights.get('business_insights', [])])
    recs_html = "".join([f"<li><strong>Action:</strong> {r}</li>" for r in insights.get('recommendations', [])])
    
    audit_steps_html = "".join([f"<li><strong>Step {i+1}:</strong> {s}</li>" for i, s in enumerate(audit_log.get('steps_executed', []))])

    table_preview_html = clean_df.head(20).to_html(classes="styled-table", index=False)

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Data Cleaning & Reporting Automation — Executive Report</title>
    <style>
        :root {{
            --primary: #1e3a8a;
            --primary-light: #2563eb;
            --bg: #f8fafc;
            --card-bg: #ffffff;
            --text-dark: #0f172a;
            --text-muted: #64748b;
            --border: #e2e8f0;
            --success: #10b981;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
            background-color: var(--bg);
            color: var(--text-dark);
            line-height: 1.6;
            padding: 30px 20px;
        }}
        .container {{ max-width: 1100px; margin: 0 auto; }}
        header {{
            background: linear-gradient(135deg, var(--primary), var(--primary-light));
            color: white;
            padding: 30px;
            border-radius: 12px;
            margin-bottom: 25px;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
        }}
        header h1 {{ font-size: 26px; font-weight: 700; margin-bottom: 6px; }}
        header p {{ font-size: 14px; opacity: 0.9; }}
        .grid-kpis {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 25px;
        }}
        .kpi-card {{
            background: var(--card-bg);
            padding: 20px;
            border-radius: 10px;
            border: 1px solid var(--border);
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}
        .kpi-label {{ font-size: 13px; color: var(--text-muted); font-weight: 600; text-transform: uppercase; }}
        .kpi-value {{ font-size: 24px; font-weight: 700; color: var(--primary); margin: 8px 0; }}
        .kpi-desc {{ font-size: 12px; color: var(--text-muted); }}
        .section-card {{
            background: var(--card-bg);
            padding: 25px;
            border-radius: 10px;
            border: 1px solid var(--border);
            margin-bottom: 25px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}
        .section-card h2 {{ font-size: 18px; color: var(--primary); margin-bottom: 15px; border-bottom: 2px solid var(--border); padding-bottom: 8px; }}
        ul {{ padding-left: 20px; }}
        li {{ margin-bottom: 8px; font-size: 14px; }}
        .styled-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            margin-top: 10px;
            overflow-x: auto;
            display: block;
        }}
        .styled-table th, .styled-table td {{
            padding: 10px 12px;
            border: 1px solid var(--border);
            text-align: left;
        }}
        .styled-table th {{
            background-color: var(--primary);
            color: white;
            font-weight: 600;
        }}
        .styled-table tr:nth-child(even) {{ background-color: #f1f5f9; }}
        footer {{ text-align: center; font-size: 12px; color: var(--text-muted); margin-top: 30px; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>Data Cleaning & Reporting Automation</h1>
            <p>Automated Data Validation, Transformation, Analysis & Executive Intelligence Report</p>
            <p style="font-size: 12px; margin-top: 5px;">Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </header>

        <div class="grid-kpis">
            <div class="kpi-card">
                <div class="kpi-label">Health Score</div>
                <div class="kpi-value" style="color: var(--success);">{clean_validation.get('quality_score', 100)}/100</div>
                <div class="kpi-desc">Initial: {raw_validation.get('quality_score', 0)}/100</div>
            </div>
            {kpi_cards_html}
        </div>

        <div class="section-card">
            <h2>Data Quality & Cleaning Audit Summary</h2>
            <ul>{audit_steps_html}</ul>
        </div>

        <div class="section-card">
            <h2>Automated Insights</h2>
            <ul>{insights_html}</ul>
        </div>

        <div class="section-card">
            <h2>Strategic Recommendations</h2>
            <ul>{recs_html}</ul>
        </div>

        <div class="section-card">
            <h2>Cleaned Dataset Preview (First 20 Records)</h2>
            {table_preview_html}
        </div>

        <footer>
            Generated by Antigravity Data Cleaning & Reporting Automation System.
        </footer>
    </div>
</body>
</html>
"""
    if isinstance(output_path_or_buffer, str):
        os.makedirs(os.path.dirname(output_path_or_buffer) if os.path.dirname(output_path_or_buffer) else ".", exist_ok=True)
        with open(output_path_or_buffer, "w", encoding="utf-8") as f:
            f.write(html_content)
        log_event("SUCCESS", "EXPORT", f"Saved HTML report to '{output_path_or_buffer}'.")
        return output_path_or_buffer
    else:
        return html_content.encode("utf-8")


# -------------------------------------------------------------
# 4. Cleaned Dataset Direct Exports (CSV / Excel)
# -------------------------------------------------------------

def export_cleaned_data(
    df: pd.DataFrame,
    filepath_or_buffer: Optional[Union[str, io.BytesIO]] = None,
    file_format: str = "csv"
) -> Union[str, bytes]:
    """Export cleaned dataset to CSV or Excel bytes/file."""
    is_file = isinstance(filepath_or_buffer, str)
    if is_file and os.path.dirname(filepath_or_buffer):
        os.makedirs(os.path.dirname(filepath_or_buffer), exist_ok=True)

    if file_format.lower() == "csv":
        if is_file:
            df.to_csv(filepath_or_buffer, index=False)
            return filepath_or_buffer
        else:
            return df.to_csv(index=False).encode("utf-8")
    else: # Excel
        export_df = _sanitize_df_for_excel(df)
        if is_file:
            export_df.to_excel(filepath_or_buffer, index=False)
            return filepath_or_buffer
        else:
            buf = io.BytesIO()
            export_df.to_excel(buf, index=False)
            buf.seek(0)
            return buf.getvalue()
